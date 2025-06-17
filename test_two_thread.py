import matplotlib.pyplot as plt
from numpy import*
import numpy as np
import pandas as pd
import openpyxl
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import copy
from threading import Thread, Lock, Event
from random import Random
from random import sample
from time import time
import inspyred
from inspyred import ec
from inspyred.ec import terminators
import sqlite3 as lite
import sys
from operator import itemgetter
import itertools
import random
from datetime import datetime
from multiprocessing import Manager 
import logging
import logging.handlers

# ============ GENE LOCATION ANALYSIS IMPORT ============
try:
    from step3_commondepath import (
        GeneCommonAncestorAnalyzer, 
        complete_gene_analysis_standalone
    )
    LOCATION_ANALYSIS_AVAILABLE = True
    print("✅ Gene location analysis module loaded successfully")
except ImportError as e:
    print(f"⚠️  Warning: Gene location analysis not available: {e}")
    LOCATION_ANALYSIS_AVAILABLE = False
    
    # Mock function for testing
    def complete_gene_analysis_standalone(g1, g2):
        """Mock location analysis function for testing"""
        return random.uniform(0, 1)

# ============ LOCATION ANALYSIS CACHE ============
location_cache = {}
cache_lock = Lock()

def cached_gene_location_analysis(g1, g2):
    """Cached wrapper for gene location analysis to avoid repeated UniProt calls"""
    # Create cache key (order-independent)
    cache_key = tuple(sorted([g1, g2]))
    
    with cache_lock:
        if cache_key in location_cache:
            return location_cache[cache_key]
    
    # If not in cache, compute the result
    try:
        result = complete_gene_analysis_standalone(g1, g2)
        
        # Store in cache
        with cache_lock:
            location_cache[cache_key] = result
            
        return result
    except Exception as e:
        print(f"⚠️  Location analysis failed for {g1}-{g2}: {e}")
        # Cache the failure to avoid repeated attempts
        with cache_lock:
            location_cache[cache_key] = 0.0
        return 0.0

# ============ GLOBAL SYNCHRONIZATION OBJECTS ============
population_lock = Lock()
start_event = Event()
shared_initial_population = None
gene_analyzer = None

if LOCATION_ANALYSIS_AVAILABLE:
    gene_analyzer = GeneCommonAncestorAnalyzer()

class ThreadSafeGA:
    """Thread-safe wrapper for genetic algorithm execution"""
    
    def __init__(self, thread_name, fitness_type, logger):
        self.thread_name = thread_name
        self.fitness_type = fitness_type  # "original" or "enhanced"
        self.logger = logger
        self.results = {}
        self.dictgene = None
        self.keyssample = None
        
    def setup_logger(self):
        """Setup thread-specific logger"""
        self.logger.propagate = False
        self.logger.setLevel(logging.DEBUG)
        
        # Create thread-specific log file
        log_filename = f'{self.thread_name}_{self.fitness_type}.log'
        file_handler = logging.handlers.RotatingFileHandler(log_filename, mode='w')
        file_handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(threadName)s - %(message)s')
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)
        
        self.logger.info(f"=== {self.fitness_type.upper()} FITNESS THREAD STARTED ===")
        
    def create_dictionnaire_final(self):
        """Create gene dictionary from CSV file"""
        l1 = list()
        l2 = list()
        
        try:
            with open('HSFinalSIVF.csv', mode='r') as infile:
                reader = csv.reader(infile, delimiter=';')
                next(reader)  # Skip header
                for rows in reader:
                    g = (rows[0], rows[1])
                    l1.append(g)
                    IS = (float(rows[2])/1000, float(rows[3]))
                    l2.append(IS)
                    
            DictF = {k: v for k, v in zip(l1, l2)}
            self.logger.info(f"Gene dictionary created with {len(DictF)} entries")
            return DictF
            
        except FileNotFoundError:
            self.logger.error("HSFinalSIVF.csv file not found!")
            return {}
            
    def get_gene_from_file(self, rand, size):
        """Generate gene community from file"""
        try:
            with open('pathW2.txt', 'r') as f2:
                l = [line.strip('\n ') for line in f2.readlines()]
                
            nv = []
            available_genes = l.copy()
            
            for i in range(size):
                if available_genes:
                    chosen_gene = rand.choice(available_genes)
                    available_genes.remove(chosen_gene)
                    nv.append(chosen_gene)
                    
            self.logger.debug(f"Generated community of size {len(nv)}: {nv}")
            return nv
            
        except FileNotFoundError:
            self.logger.error("pathW2.txt file not found!")
            return []

    def generate_pop(self, random, args):
        """Generate population using shared initial population"""
        global shared_initial_population
        
        with population_lock:
            if shared_initial_population is None:
                # First thread generates the population
                tabmax = 40
                size = random.randint(5, tabmax)
                shared_initial_population = self.get_gene_from_file(random, size)
                self.logger.info(f"Created shared initial population: {shared_initial_population}")
            else:
                self.logger.info(f"Using shared initial population: {shared_initial_population}")
                
        return shared_initial_population.copy()

    def fitness_original(self, individu, dictgene):
        """Original fitness function: similarity + interaction only"""
        st = 0  # total pairs
        si = 0.0  # sum interactions
        ss = 0.0  # sum similarities
        alpha = 0.5
        beta = 0.5
        
        for i, item in enumerate(individu):
            for j in range(i+1, len(individu)):
                g1 = individu[i]
                g2 = individu[j]
                genes = (g1, g2)
                geness = (g2, g1)
                t = len(individu)
                
                v1 = dictgene.get(genes, 0)
                v2 = dictgene.get(geness, 0)
                
                if v1 == 0:
                    v = v2
                else:
                    v = v1
                    
                if v != 0:
                    intr = v[0]
                    s = v[1]
                    si = si + intr
                    ss = ss + s
                else:
                    intr = 0.0
                    s = 0.0
                    
            st = st + (t-i)
            
        sint = si/st if st > 0 else 0.0
        smoyen = ss/st if st > 0 else 0.0
        f = alpha*smoyen + beta*sint
        
        return sint, smoyen, f

    def fitness_enhanced(self, individu, dictgene):
        """Enhanced fitness function: similarity + interaction + location"""
        st = 0  # total pairs
        si = 0.0  # sum interactions
        ss = 0.0  # sum similarities
        sl = 0.0  # sum location scores
        location_pairs = 0
        
        alpha = 0.4  # interactions
        beta = 0.4   # similarities
        gamma = 0.2  # location scores
        
        for i, item in enumerate(individu):
            for j in range(i+1, len(individu)):
                g1 = individu[i]
                g2 = individu[j]
                genes = (g1, g2)
                geness = (g2, g1)
                t = len(individu)
                
                v1 = dictgene.get(genes, 0)
                v2 = dictgene.get(geness, 0)
                
                if v1 == 0:
                    v = v2
                else:
                    v = v1
                    
                if v != 0:
                    intr = v[0]
                    s = v[1]
                    si = si + intr
                    ss = ss + s
                else:
                    intr = 0.0
                    s = 0.0
                
                # Calculate location score
                try:
                    location_score = cached_gene_location_analysis(g1, g2)  # CHANGED: Use cached version
                    if location_score > 0:
                        sl += location_score
                        location_pairs += 1
                except Exception as e:
                    self.logger.warning(f"Location analysis failed for {g1}-{g2}: {e}")
                    location_score = 0.0
                    
            st = st + (t-i)
            
        sint = si/st if st > 0 else 0.0
        smoyen = ss/st if st > 0 else 0.0
        
        if location_pairs > 0:
            slocation = sl/location_pairs
            f = alpha*smoyen + beta*sint + gamma*slocation
            self.logger.debug(f"Location pairs: {location_pairs}/{st//2}, Avg location: {slocation:.3f}")
        else:
            slocation = 0.0
            # Fallback to original weights when no location data
            alpha, beta = 0.5, 0.5
            f = alpha*smoyen + beta*sint
            self.logger.debug("No location data available, using interaction+similarity only")
            
        return sint, smoyen, slocation, f

    def fitness(self, candidates, args):
        """Main fitness evaluation function"""
        self.logger.debug("=== FITNESS EVALUATION ===")
        l_fit = []
        
        for i, individu in enumerate(candidates):
            if self.fitness_type == "original":
                sint, smoyen, f = self.fitness_original(individu, self.dictgene)
                self.logger.debug(f'Individual {i+1}: Interaction={sint:.3f}, Similarity={smoyen:.3f}, Fitness={f:.3f}')
            else:  # enhanced
                sint, smoyen, slocation, f = self.fitness_enhanced(individu, self.dictgene)
                self.logger.debug(f'Individual {i+1}: Interaction={sint:.3f}, Similarity={smoyen:.3f}, Location={slocation:.3f}, Fitness={f:.3f}')
                
            l_fit.append(f)
            
        self.logger.debug(f"Fitness list: {l_fit}")
        return l_fit

    def get_best_worst_gene(self, candidate):
        """Get best and worst genes in a candidate"""
        gene_scores = []
        
        for idx, gene in enumerate(candidate):
            si = 0.0
            ss = 0.0
            
            for j, other_gene in enumerate(candidate):
                if idx != j:
                    genes = (gene, other_gene)
                    geness = (other_gene, gene)
                    
                    v1 = self.dictgene.get(genes, 0)
                    v2 = self.dictgene.get(geness, 0)
                    
                    if v1 == 0:
                        v = v2
                    else:
                        v = v1
                        
                    if v != 0:
                        si += float(v[0])
                        ss += float(v[1])
                        
            t = len(candidate)
            avg_sim = ss / (t-1) if t > 1 else 0
            avg_int = si / (t-1) if t > 1 else 0
            score = avg_sim + avg_int
            gene_scores.append((gene, score))
            
        if gene_scores:
            best_gene = max(gene_scores, key=itemgetter(1))
            worst_gene = min(gene_scores, key=itemgetter(1))
            return best_gene, worst_gene
        return None, None

    def get_gene_interaction(self, mutant, best_gene, rand):
        """Get genes that interact with best gene"""
        interacting_genes = []
        
        for k in self.keyssample:
            if best_gene == k[0]:
                interacting_genes.append(k[1])
            if best_gene == k[1]:
                interacting_genes.append(k[0])
                
        # Remove duplicates and genes already in mutant
        interacting_genes = list(set(interacting_genes))
        available_genes = [g for g in interacting_genes if g not in mutant]
        
        if not available_genes:
            # If no interacting genes available, choose random gene from mutant
            choice = rand.choice([gene for gene in mutant if gene != best_gene])
        else:
            choice = rand.choice(available_genes)
            
        self.logger.debug(f"Gene interaction choice: {choice}")
        return choice

    def mutate_gene(self, random, child, args):
        """Custom mutation function"""
        self.logger.debug("=== MUTATION ===")
        threshold = 0.5
        
        mutant = list(set(child))  # Remove duplicates
        self.logger.debug(f"Mutant before mutation: {mutant} (size: {len(mutant)})")
        
        # Get gene scores
        best_gene, worst_gene = self.get_best_worst_gene(mutant)
        
        if best_gene and worst_gene:
            self.logger.debug(f"Best gene: {best_gene[0]} (score: {best_gene[1]:.3f})")
            self.logger.debug(f"Worst gene: {worst_gene[0]} (score: {worst_gene[1]:.3f})")
            
            gene_to_insert = self.get_gene_interaction(mutant, best_gene[0], random)
            
            if worst_gene[1] >= threshold:
                # Insertion
                self.logger.debug("***** INSERTION *****")
                mutant.append(gene_to_insert)
            else:
                # Replacement
                self.logger.debug("***** REPLACEMENT *****")
                for k, gene in enumerate(mutant):
                    if gene == worst_gene[0]:
                        mutant[k] = gene_to_insert
                        break
                        
        candidate = copy.copy(mutant)
        self.logger.debug(f"Mutant after mutation: {candidate}")
        return candidate

    def test_observers(self, population, num_generations, num_evaluations, args):
        """Custom observer for logging"""
        self.logger.info(f"=== GENERATION {num_generations} ===")
        self.logger.info(f"Population size: {len(population)}")
        
        for i, individual in enumerate(population):
            self.logger.debug(f"Individual {i+1}: {individual}")
            
        # Call standard observers
        inspyred.ec.observers.stats_observer(population, num_generations, num_evaluations, args)

    def get_infos(self, size, best_fit, duration, best_candidate):
        """Generate DataFrame with results"""
        if self.fitness_type == "original":
            sint, smoyen, f = self.fitness_original(best_candidate, self.dictgene)
            df = pd.DataFrame([[size, '', best_fit, duration, sint, smoyen, str(best_candidate)]], 
                            columns=['Size', '% Coverage', 'Fitness', 'Duration', 'Avg_Int', 'Avg_Sim', 'Individual'])
        else:  # enhanced
            sint, smoyen, slocation, f = self.fitness_enhanced(best_candidate, self.dictgene)
            df = pd.DataFrame([[size, '', best_fit, duration, sint, smoyen, slocation, str(best_candidate)]], 
                            columns=['Size', '% Coverage', 'Fitness', 'Duration', 'Avg_Int', 'Avg_Sim', 'Avg_Loc', 'Individual'])
        return df

    def generate_file(self, file, df):
        """Generate/append Excel file"""
        sheet_name = f'results_{self.fitness_type}'
        
        if os.path.isfile(file):
            workbook = openpyxl.load_workbook(file)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                # Add headers
                for col, header in enumerate(df.columns, 1):
                    sheet.cell(row=1, column=col, value=header)
            
            # Append data
            for row in dataframe_to_rows(df, header=False, index=False):
                sheet.append(row)
            workbook.save(file)
            workbook.close()
        else:
            with pd.ExcelWriter(path=file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
        self.logger.info(f"Results saved to {file}")

    def run_algorithm(self):
        """Main algorithm execution"""
        try:
            start_time = datetime.now()
            self.logger.info(f"Algorithm started at {start_time}")
            
            # Wait for all threads to be ready
            start_event.wait()
            
            # Setup data
            self.dictgene = self.create_dictionnaire_final()
            self.keyssample = self.dictgene.keys()
            
            # Setup random seed for reproducibility within thread
            rand = Random()
            rand.seed(42 + hash(self.thread_name))  # Different but deterministic seeds
            
            # Create algorithm
            algorithm = inspyred.ec.GA(rand)
            algorithm.terminator = inspyred.ec.terminators.generation_termination
            algorithm.observer = self.test_observers
            algorithm.selector = inspyred.ec.selectors.tournament_selection
            algorithm.replacer = inspyred.ec.replacers.steady_state_replacement
            algorithm.variator = [inspyred.ec.variators.n_point_crossover, 
                                inspyred.ec.variators.mutator(self.mutate_gene)]
            algorithm.archiver = inspyred.ec.archivers.population_archiver
            algorithm.migrator = inspyred.ec.migrators.default_migration
            
            # Run evolution
            self.logger.info("Starting evolution...")
            final_pop = algorithm.evolve(
                generator=self.generate_pop,
                evaluator=self.fitness,
                pop_size=10,              # Reduced for testing
                maximize=True,
                num_selected=2,
                tournament_size=2,
                max_evaluations=50,       # Reduced for testing
                mutation_rate=0.1,
                crossover_rate=0.8,
                max_generations=10        # Reduced for testing
            )
            
            # Process results
            end_time = datetime.now()
            duration = str(end_time - start_time)
            
            if final_pop:
                best = max(final_pop)
                self.logger.info(f"Evolution completed successfully")
                self.logger.info(f"Best solution: {best}")
                self.logger.info(f"Duration: {duration}")
                
                # Save results
                df = self.get_infos(len(best.candidate), best.fitness, duration, best.candidate)
                filename = f'results_{self.fitness_type}.xlsx'
                self.generate_file(filename, df)
                
                # Store results for comparison
                self.results = {
                    'best_fitness': best.fitness,
                    'best_candidate': best.candidate,
                    'duration': duration,
                    'population_size': len(final_pop)
                }
                
            else:
                self.logger.error("Evolution failed - no final population")
                
        except Exception as e:
            self.logger.error(f"Algorithm execution failed: {e}", exc_info=True)

def run_thread(thread_name, fitness_type):
    """Thread execution function"""
    # Create thread-specific logger
    logger = logging.getLogger(f'ga_{fitness_type}')
    
    # Create and run algorithm
    ga_instance = ThreadSafeGA(thread_name, fitness_type, logger)
    ga_instance.setup_logger()
    ga_instance.run_algorithm()
    
    return ga_instance.results

def main():
    """Main execution function"""
    print("🧬 Starting Threaded Genetic Algorithm for Gene Community Analysis")
    print("=" * 70)
    
    global shared_initial_population
    shared_initial_population = None
    
    # Clear location cache at start
    global location_cache
    location_cache.clear()
    print("🗄️  Location analysis cache initialized")
    
    # Create threads
    thread1 = Thread(target=run_thread, args=("Thread-1", "original"), name="OriginalGA")
    thread2 = Thread(target=run_thread, args=("Thread-2", "enhanced"), name="EnhancedGA")
    
    print("🚀 Starting both threads...")
    
    # Start threads
    thread1.start()
    thread2.start()
    
    # Signal threads to start algorithm execution
    start_event.set()
    
    print("⏳ Algorithms running in parallel...")
    print("📊 Check log files for detailed progress:")
    print("   - Thread-1_original.log")
    print("   - Thread-2_enhanced.log")
    print("📁 Results will be saved to:")
    print("   - results_original.xlsx")
    print("   - results_enhanced.xlsx")
    print("💡 Location analysis results are cached to avoid repeated UniProt calls")
    
    # Wait for completion
    thread1.join()
    thread2.join()
    
    print(f"\n✅ Both algorithms completed!")
    print(f"🗄️  Location cache contains {len(location_cache)} gene pair results")
    print("🔍 Check the log files and Excel outputs for detailed analysis.")
    print("=" * 70)

if __name__ == '__main__':
    main()